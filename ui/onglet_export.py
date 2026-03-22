"""
ui/onglet_export.py
===================
OngletExport — Export Excel synthèse AM + planning hebdomadaire

Feuille 1 : Synthèse par employé (Nb AM / Absences / AM nets) par année
Feuille 2 : Planning hebdomadaire semaine par semaine
"""

import os
from datetime import date as _date, timedelta, datetime
from collections import defaultdict

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QFileDialog, QMessageBox, QProgressBar, QTextEdit,
    QGroupBox, QDateEdit, QCheckBox,
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QDate
from PyQt6.QtGui import QFont

from ui.fermetures import fusionner_absences_fermetures, jours_fermetures_periode
from ui.constantes import (
    COULEURS,
    PLANNING_HISTORIQUE_JSON, EMPLOYES_CONTRATS_JSON,
    ABSENCES_JSON,
    charger_json,
)


# ═══════════════════════════════════════════════════════════════
# LOGIQUE MÉTIER — calculs AM / absences
# ═══════════════════════════════════════════════════════════════

ANNEES = [2021, 2022, 2023, 2024, 2025]


def _parse_date_jj_mm_aaaa(s: str):
    """'JJ-MM-AAAA' → date. None si invalide."""
    if not s:
        return None
    try:
        j, m, a = s.split("-")
        return _date(int(a), int(m), int(j))
    except Exception:
        return None


def _jours_ouvres_plage(d_debut: _date, d_fin: _date) -> list:
    """Retourne la liste des jours ouvrés (lun-ven) entre d_debut et d_fin."""
    jours = []
    d = d_debut
    while d <= d_fin:
        if d.weekday() < 5:
            jours.append(d)
        d += timedelta(days=1)
    return jours


def _construire_cache_absences(absences_data: dict) -> dict:
    """
    Construit un cache {cle_emp: set(dates absences ouvrées)}.
    """
    cache = {}
    for cle_emp, periodes in absences_data.items():
        if not isinstance(periodes, list):
            continue
        jours_abs = set()
        for p in periodes:
            d = _parse_date_jj_mm_aaaa(p.get("debut", ""))
            f = _parse_date_jj_mm_aaaa(p.get("fin", ""))
            if not d or not f:
                continue
            for j in _jours_ouvres_plage(d, f):
                jours_abs.add(j)
        cache[cle_emp] = jours_abs
    return cache


def _synthese_semaine_depuis_jours(jours_dict: dict, cle_sem: str) -> tuple:
    """
    Retourne (cycle_dominant, hypothetique) pour une semaine ISO
    en synthétisant les données jours.
    Priorité : jours réels > jours hypothétiques.
    """
    from collections import Counter
    try:
        parts = cle_sem.split('_')
        annee = int(parts[1])
        num   = int(parts[0][1:])
        lundi = datetime.strptime(f"{annee}-W{num:02d}-1", "%G-W%V-%u").date()
    except Exception:
        return None, None

    cycles_reels = []
    cycles_hyp   = []
    for offset in range(5):  # lun-ven seulement
        jour = lundi + timedelta(days=offset)
        cle_j = jour.strftime("%Y-%m-%d")
        entree = jours_dict.get(cle_j)
        if not entree:
            continue
        c   = entree.get('cycle')
        hyp = entree.get('hypothetique', False)
        if c and c != 'R':
            if not hyp:
                cycles_reels.append(c)
            else:
                cycles_hyp.append(c)

    if cycles_reels:
        dominant = Counter(cycles_reels).most_common(1)[0][0]
        return dominant, False
    if cycles_hyp:
        dominant = Counter(cycles_hyp).most_common(1)[0][0]
        return dominant, True
    return None, None


def _get_cycle_semaine(data_emp: dict, cle_sem: str) -> tuple:
    """
    Retourne (cycle, hypothetique) pour une semaine.
    Priorité : jours réels > semaines réelles > jours hyp > semaines hyp.
    """
    jours_dict = data_emp.get('jours', {})
    semaines_dict = data_emp.get('semaines', {})

    # 1. Synthèse depuis jours réels
    cycle, hyp = _synthese_semaine_depuis_jours(jours_dict, cle_sem)
    if cycle is not None and not hyp:
        return cycle, False

    # 2. Semaine réelle dans 'semaines'
    entree_sem = semaines_dict.get(cle_sem)
    if entree_sem and not entree_sem.get('hypothetique', True):
        return entree_sem.get('cycle'), False

    # 3. Jours hypothétiques
    if cycle is not None:
        return cycle, True

    # 4. Semaine hypothétique
    if entree_sem:
        return entree_sem.get('cycle'), entree_sem.get('hypothetique', True)

    return None, None


def calculer_stats_employe(
    cle_emp: str,
    data_emp: dict,
    absences_emp: set,
    d_debut: _date,
    d_fin: _date,
) -> dict:
    """
    Calcule les statistiques AM pour un employé sur la période.
    Retourne un dict avec les stats par année et totaux.
    """
    stats = {}

    # Collecter les jours AM (réels ET hypothétiques) sur la période
    jours_am_tous = set()
    for cle_j, val in data_emp.get('jours', {}).items():
        if val.get('cycle') == 'AM':  # réels + hypothétiques
            try:
                d = _date.fromisoformat(cle_j)
                if d_debut <= d <= d_fin:
                    jours_am_tous.add(d)
            except Exception:
                pass

    if not jours_am_tous:
        return {}  # Pas de jours AM → exclure de l'export

    for annee in ANNEES:
        d_debut_an = max(d_debut, _date(annee, 1, 1))
        d_fin_an   = min(d_fin,   _date(annee, 12, 31))
        if d_debut_an > d_fin_an:
            stats[annee] = {"abs_total": 0, "abs_am": 0, "jours_am": 0, "am_nets": 0}
            continue

        # Jours AM (réels + hypothétiques) cette année
        am_annee = {d for d in jours_am_tous if d.year == annee}

        # Absences ouvrées cette année (personnelles + fermetures obligatoires, sans doublon)
        fermetures_annee = jours_fermetures_periode(d_debut_an, d_fin_an)
        abs_annee = {d for d in (absences_emp | fermetures_annee)
                     if d_debut_an <= d <= d_fin_an}

        # Absences tombant sur un jour AM
        abs_sur_am = abs_annee & am_annee

        stats[annee] = {
            "abs_total": len(abs_annee),
            "abs_am":    len(abs_sur_am),
            "jours_am":  len(am_annee),
            "am_nets":   len(am_annee) - len(abs_sur_am),
        }

    return stats


# ═══════════════════════════════════════════════════════════════
# WORKER THREAD — génération Excel en arrière-plan
# ═══════════════════════════════════════════════════════════════

class WorkerExport(QThread):
    log_signal    = pyqtSignal(str)
    fini_signal   = pyqtSignal(str)   # chemin fichier généré
    erreur_signal = pyqtSignal(str)

    def __init__(self, chemin_sortie: str, d_debut: _date, d_fin: _date,
                 inclure_feuille2: bool):
        super().__init__()
        self.chemin_sortie    = chemin_sortie
        self.d_debut          = d_debut
        self.d_fin            = d_fin
        self.inclure_feuille2 = inclure_feuille2

    def run(self):
        try:
            import openpyxl
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side
            )
            from openpyxl.utils import get_column_letter
        except ImportError:
            self.erreur_signal.emit("openpyxl est requis. Installez-le avec : pip install openpyxl")
            return

        def log(msg):
            self.log_signal.emit(msg)

        try:
            log("📂  Chargement des données…")
            planning  = charger_json(PLANNING_HISTORIQUE_JSON())
            employes  = charger_json(EMPLOYES_CONTRATS_JSON())
            absences  = charger_json(ABSENCES_JSON())

            log("🔢  Calcul des absences…")
            cache_abs = _construire_cache_absences(absences)

            # Calculer stats pour tous les employés
            log("📊  Calcul des statistiques AM…")
            resultats = {}
            for cle_emp, data_emp in planning.items():
                if cle_emp == "COMMENTAIRE" or "|" not in cle_emp:
                    continue
                absences_emp = cache_abs.get(cle_emp, set())
                stats = calculer_stats_employe(
                    cle_emp, data_emp, absences_emp,
                    self.d_debut, self.d_fin
                )
                if stats:
                    resultats[cle_emp] = stats

            log(f"   {len(resultats)} employé(s) avec jours AM réels")

            # Créer le classeur
            wb = openpyxl.Workbook()

            # ─────────────────────────────────────────────
            # STYLES
            # ─────────────────────────────────────────────
            FONT_TITRE    = Font(name="Arial", bold=True, size=11, color="FFFFFF")
            FONT_ENTETE   = Font(name="Arial", bold=True, size=10)
            FONT_NORMAL   = Font(name="Arial", size=10)
            FONT_TOTAL    = Font(name="Arial", bold=True, size=10)

            FILL_TITRE    = PatternFill("solid", start_color="2E4057")
            FILL_ANNEE    = PatternFill("solid", start_color="4A90D9")
            FILL_TOTAL    = PatternFill("solid", start_color="D9E8F5")
            FILL_AM       = PatternFill("solid", start_color="C8E6C9")  # vert clair
            FILL_VIDE     = PatternFill("solid", start_color="F5F5F5")
            FILL_HYP      = PatternFill("solid", start_color="FFF9C4")  # jaune clair hyp

            ALIGN_CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ALIGN_LEFT    = Alignment(horizontal="left",   vertical="center")
            ALIGN_RIGHT   = Alignment(horizontal="right",  vertical="center")

            BORD_FIN = Side(style="thin", color="BBBBBB")
            BORD_MED = Side(style="medium", color="888888")

            def bord(top=None, bottom=None, left=None, right=None):
                return Border(top=top, bottom=bottom, left=left, right=right)

            def style_cell(cell, font=None, fill=None, align=None, border=None, fmt=None):
                if font:   cell.font      = font
                if fill:   cell.fill      = fill
                if align:  cell.alignment = align
                if border: cell.border    = border
                if fmt:    cell.number_format = fmt

            # ─────────────────────────────────────────────
            # FEUILLE 1 — Synthèse AM
            # ─────────────────────────────────────────────
            log("📝  Génération feuille 1 — Synthèse AM…")
            ws1 = wb.active
            ws1.title = "Synthèse AM"
            ws1.freeze_panes = "D3"

            # Ligne 1 : titre principal
            ws1.merge_cells("A1:C1")
            ws1["A1"] = "SYNTHÈSE CYCLES APRÈS-MIDI"
            style_cell(ws1["A1"], font=Font(name="Arial", bold=True, size=13, color="FFFFFF"),
                       fill=FILL_TITRE, align=ALIGN_CENTER)
            periode_txt = f"{self.d_debut.strftime('%d/%m/%Y')} → {self.d_fin.strftime('%d/%m/%Y')}"
            # Colonnes fixes : Nom / Prénom / Matricule
            # Puis par année : Abs Total / Abs AM / Jours AM / AM Nets
            # Puis Totaux
            nb_cols_fixes = 5  # Nom, Prénom, Matricule, Date début, Date fin
            nb_cols_annee = 4  # Abs Total, Abs AM, Jours AM, AM Nets
            nb_annees     = len(ANNEES)
            nb_cols_total = nb_cols_fixes + nb_cols_annee * nb_annees + nb_cols_annee

            # Titre période sur toute la largeur
            derniere_col = get_column_letter(nb_cols_total)
            ws1.merge_cells(f"D1:{derniere_col}1")
            ws1["D1"] = f"Période : {periode_txt}"
            style_cell(ws1["D1"], font=Font(name="Arial", bold=True, size=11, color="FFFFFF"),
                       fill=FILL_TITRE, align=ALIGN_CENTER)

            # Ligne 2 : en-têtes colonnes fixes
            entetes_fixes = ["Nom", "Prénom", "Matricule", "Date début", "Date fin"]
            for col_idx, titre in enumerate(entetes_fixes, start=1):
                cell = ws1.cell(row=2, column=col_idx, value=titre)
                style_cell(cell, font=FONT_ENTETE, fill=FILL_ANNEE,
                           align=ALIGN_CENTER,
                           border=bord(bottom=BORD_MED))
                cell.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")

            # Groupes par année
            col_start = nb_cols_fixes + 1
            for annee in ANNEES:
                col_fin = col_start + nb_cols_annee - 1
                # Fusionner entête année
                ws1.merge_cells(
                    start_row=2, start_column=col_start,
                    end_row=2,   end_column=col_fin
                )
                cell_an = ws1.cell(row=2, column=col_start, value=str(annee))
                style_cell(cell_an,
                           font=Font(name="Arial", bold=True, size=10, color="FFFFFF"),
                           fill=FILL_ANNEE, align=ALIGN_CENTER,
                           border=bord(bottom=BORD_MED, left=BORD_MED))
                col_start += nb_cols_annee

            # Totaux
            ws1.merge_cells(
                start_row=2, start_column=col_start,
                end_row=2,   end_column=col_start + nb_cols_annee - 1
            )
            cell_tot = ws1.cell(row=2, column=col_start, value="TOTAL")
            style_cell(cell_tot,
                       font=Font(name="Arial", bold=True, size=10, color="FFFFFF"),
                       fill=PatternFill("solid", start_color="1A3A5C"),
                       align=ALIGN_CENTER,
                       border=bord(bottom=BORD_MED, left=BORD_MED))

            # Ligne 3 : sous-entêtes
            sous_entetes = ["Abs Total", "Abs AM", "Jours AM", "AM Nets"]
            col = nb_cols_fixes + 1
            for _ in range(nb_annees + 1):  # +1 pour totaux
                for se in sous_entetes:
                    cell = ws1.cell(row=3, column=col, value=se)
                    style_cell(cell, font=Font(name="Arial", bold=True, size=9),
                               fill=FILL_VIDE, align=ALIGN_CENTER,
                               border=bord(bottom=BORD_FIN))
                    col += 1

            # Sous-entêtes colonnes fixes (ligne 3 = fusion avec ligne 2)
            for c in range(1, nb_cols_fixes + 1):
                ws1.merge_cells(start_row=2, start_column=c, end_row=3, end_column=c)

            ws1.row_dimensions[1].height = 24
            ws1.row_dimensions[2].height = 22
            ws1.row_dimensions[3].height = 22

            # Données
            row_data = 4
            for cle_emp in sorted(resultats.keys()):
                stats = resultats[cle_emp]
                info_emp = employes.get(cle_emp, {})

                # Parser nom/prénom/matricule depuis la clé
                partie_nom, matricule = cle_emp.split("|", 1)
                parties = partie_nom.strip().split(" ", 1)
                nom    = parties[0] if parties else partie_nom
                prenom = parties[1] if len(parties) > 1 else ""

                # Date début contrat
                date_debut_contrat = info_emp.get("date_debut", "")

                # Colonnes fixes
                vals_fixes = [nom, prenom, matricule,
                              self.d_debut.strftime("%d/%m/%Y"),
                              self.d_fin.strftime("%d/%m/%Y")]
                for c_idx, val in enumerate(vals_fixes, start=1):
                    cell = ws1.cell(row=row_data, column=c_idx, value=val)
                    style_cell(cell, font=FONT_NORMAL, align=ALIGN_LEFT)

                # Stats par année
                col = nb_cols_fixes + 1
                totaux = {"abs_total": 0, "abs_am": 0, "jours_am": 0, "am_nets": 0}
                for annee in ANNEES:
                    s = stats.get(annee, {"abs_total": 0, "abs_am": 0, "jours_am": 0, "am_nets": 0})
                    for cle_s in ["abs_total", "abs_am", "jours_am", "am_nets"]:
                        val = s[cle_s]
                        cell = ws1.cell(row=row_data, column=col, value=val if val else "")
                        fill = FILL_AM if cle_s == "jours_am" and val > 0 else None
                        style_cell(cell, font=FONT_NORMAL, align=ALIGN_CENTER, fill=fill)
                        totaux[cle_s] += val
                        col += 1

                # Totaux ligne
                for cle_s in ["abs_total", "abs_am", "jours_am", "am_nets"]:
                    val = totaux[cle_s]
                    cell = ws1.cell(row=row_data, column=col, value=val if val else "")
                    fill = PatternFill("solid", start_color="A5D6A7") if cle_s == "jours_am" and val > 0 else FILL_TOTAL
                    style_cell(cell, font=FONT_TOTAL, align=ALIGN_CENTER, fill=fill)
                    col += 1

                row_data += 1

            # Ligne totaux globaux
            ws1.cell(row=row_data, column=1, value="TOTAL GLOBAL").font = FONT_TOTAL
            ws1.cell(row=row_data, column=1).fill = FILL_TOTAL
            col = nb_cols_fixes + 1
            for _ in range(nb_annees + 1):
                for _ in sous_entetes:
                    r_start = 4
                    r_end   = row_data - 1
                    col_letter = get_column_letter(col)
                    cell = ws1.cell(row=row_data, column=col,
                                    value=f"=SUM({col_letter}{r_start}:{col_letter}{r_end})")
                    style_cell(cell, font=FONT_TOTAL, align=ALIGN_CENTER,
                               fill=FILL_TOTAL, fmt="0")
                    col += 1

            # Largeurs colonnes feuille 1
            ws1.column_dimensions["A"].width = 20
            ws1.column_dimensions["B"].width = 16
            ws1.column_dimensions["C"].width = 12
            ws1.column_dimensions["D"].width = 13
            ws1.column_dimensions["E"].width = 13
            for c_idx in range(nb_cols_fixes + 1, nb_cols_total + 1):
                ws1.column_dimensions[get_column_letter(c_idx)].width = 10

            log(f"   ✅ Feuille 1 : {row_data - 4} lignes")

            # ─────────────────────────────────────────────
            # FEUILLE 2 — Planning hebdomadaire
            # ─────────────────────────────────────────────
            if self.inclure_feuille2:
                log("📅  Génération feuille 2 — Planning hebdomadaire…")
                ws2 = wb.create_sheet("Planning hebdomadaire")
                ws2.freeze_panes = "D2"

                # Générer la liste des semaines ISO de la période
                semaines = []
                d = self.d_debut - timedelta(days=self.d_debut.weekday())
                while d <= self.d_fin:
                    iso = d.isocalendar()
                    cle_sem = f"S{iso[1]:02d}_{iso[0]}"
                    semaines.append((cle_sem, d))
                    d += timedelta(weeks=1)

                # En-tête ligne 1
                ws2.cell(row=1, column=1, value="Nom").font        = FONT_ENTETE
                ws2.cell(row=1, column=2, value="Prénom").font     = FONT_ENTETE
                ws2.cell(row=1, column=3, value="Matricule").font  = FONT_ENTETE
                for c_idx, (cle_sem, lundi) in enumerate(semaines, start=4):
                    cell = ws2.cell(row=1, column=c_idx,
                                    value=f"{cle_sem}\n{lundi.strftime('%d/%m')}")
                    style_cell(cell,
                               font=Font(name="Arial", bold=True, size=8),
                               fill=FILL_ANNEE, align=ALIGN_CENTER)
                    cell.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")

                for c in range(1, 4):
                    ws2.cell(row=1, column=c).fill = FILL_TITRE
                    ws2.cell(row=1, column=c).font = Font(name="Arial", bold=True,
                                                          size=10, color="FFFFFF")
                    ws2.cell(row=1, column=c).alignment = ALIGN_CENTER

                # Couleurs cycles
                COULEURS_CYCLES_XL = {
                    "AM": ("2E7D32", "FFFFFF"),   # vert foncé
                    "M":  ("E65100", "FFFFFF"),   # orange foncé
                    "N":  ("4A148C", "FFFFFF"),   # violet
                    "J":  ("546E7A", "FFFFFF"),   # gris bleu
                    "WE": ("1565C0", "FFFFFF"),   # bleu
                    "R":  ("EEEEEE", "333333"),   # gris clair
                }
                COULEURS_CYCLES_HYP = {
                    "AM": ("A5D6A7", "1B5E20"),
                    "M":  ("FFCC80", "BF360C"),
                    "N":  ("CE93D8", "4A148C"),
                    "J":  ("B0BEC5", "37474F"),
                    "WE": ("90CAF9", "0D47A1"),
                    "R":  ("F5F5F5", "757575"),
                }

                # Données
                row_p = 2
                cles_a_exporter = sorted(resultats.keys())
                for cle_emp in cles_a_exporter:
                    data_emp = planning.get(cle_emp, {})
                    partie_nom, matricule = cle_emp.split("|", 1)
                    parties = partie_nom.strip().split(" ", 1)
                    nom    = parties[0] if parties else partie_nom
                    prenom = parties[1] if len(parties) > 1 else ""

                    ws2.cell(row=row_p, column=1, value=nom).font    = FONT_NORMAL
                    ws2.cell(row=row_p, column=2, value=prenom).font = FONT_NORMAL
                    ws2.cell(row=row_p, column=3, value=matricule).font = FONT_NORMAL
                    for c in range(1, 4):
                        ws2.cell(row=row_p, column=c).alignment = ALIGN_LEFT

                    for c_idx, (cle_sem, _) in enumerate(semaines, start=4):
                        cycle, hyp = _get_cycle_semaine(data_emp, cle_sem)
                        if cycle:
                            cell = ws2.cell(row=row_p, column=c_idx, value=cycle)
                            palette = COULEURS_CYCLES_HYP if hyp else COULEURS_CYCLES_XL
                            bg, fg = palette.get(cycle, ("FFFFFF", "000000"))
                            cell.fill      = PatternFill("solid", start_color=bg)
                            cell.font      = Font(name="Arial", size=9,
                                                  color=fg,
                                                  italic=hyp)
                            cell.alignment = ALIGN_CENTER
                        else:
                            ws2.cell(row=row_p, column=c_idx, value="")

                    row_p += 1

                # Largeurs feuille 2
                ws2.column_dimensions["A"].width = 20
                ws2.column_dimensions["B"].width = 16
                ws2.column_dimensions["C"].width = 12
                for c_idx in range(4, len(semaines) + 4):
                    ws2.column_dimensions[get_column_letter(c_idx)].width = 8

                ws2.row_dimensions[1].height = 32

                log(f"   ✅ Feuille 2 : {row_p - 2} employés × {len(semaines)} semaines")

            log("💾  Sauvegarde du fichier…")
            wb.save(self.chemin_sortie)
            log(f"✅  Export terminé : {os.path.basename(self.chemin_sortie)}")
            self.fini_signal.emit(self.chemin_sortie)

        except Exception as e:
            import traceback
            self.erreur_signal.emit(f"{e}\n\n{traceback.format_exc()}")


# ═══════════════════════════════════════════════════════════════
# ONGLET EXPORT
# ═══════════════════════════════════════════════════════════════

class OngletExport(QWidget):
    """Onglet génération du fichier Excel de synthèse AM."""

    def __init__(self):
        super().__init__()
        self._worker = None
        self._construire_ui()

    def _construire_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(14)
        layout.setContentsMargins(16, 16, 16, 16)

        # Titre
        titre = QLabel("📤  Export Excel — Synthèse Cycles AM")
        titre.setStyleSheet(
            f"font-size: 18px; font-weight: 700; color: {COULEURS['texte']};"
        )
        layout.addWidget(titre)

        # ── Groupe paramètres ────────────────────────────────
        grp_params = QGroupBox("Paramètres de la synthèse")
        grp_params.setStyleSheet(
            f"QGroupBox {{ color: {COULEURS['accent']}; "
            f"border: 1px solid {COULEURS['bordure']}; "
            f"border-radius: 6px; margin-top: 6px; padding: 12px; font-size: 12px; }}"
        )
        lay_params = QVBoxLayout(grp_params)

        # Dates
        row_dates = QHBoxLayout()
        lbl_debut = QLabel("Date début :")
        lbl_debut.setStyleSheet(f"color: {COULEURS['texte_secondaire']}; font-size: 12px;")
        self.date_debut = QDateEdit()
        self.date_debut.setCalendarPopup(True)
        self.date_debut.setDate(QDate(2021, 1, 1))
        self.date_debut.setDisplayFormat("dd/MM/yyyy")
        self.date_debut.setFixedWidth(130)

        lbl_fin = QLabel("Date fin :")
        lbl_fin.setStyleSheet(f"color: {COULEURS['texte_secondaire']}; font-size: 12px;")
        self.date_fin = QDateEdit()
        self.date_fin.setCalendarPopup(True)
        self.date_fin.setDate(QDate.currentDate())
        self.date_fin.setDisplayFormat("dd/MM/yyyy")
        self.date_fin.setFixedWidth(130)

        for w in [lbl_debut, self.date_debut, lbl_fin, self.date_fin]:
            row_dates.addWidget(w)
        row_dates.addStretch()
        lay_params.addLayout(row_dates)

        # Option feuille 2
        self.check_feuille2 = QCheckBox("Inclure la feuille Planning hebdomadaire (Feuille 2)")
        self.check_feuille2.setChecked(True)
        self.check_feuille2.setStyleSheet(
            f"color: {COULEURS['texte']}; font-size: 12px;"
        )
        lay_params.addWidget(self.check_feuille2)

        layout.addWidget(grp_params)

        # ── Bouton export ────────────────────────────────────
        row_btn = QHBoxLayout()
        self.btn_exporter = QPushButton("📤  Générer l'export Excel")
        self.btn_exporter.setFixedHeight(40)
        self.btn_exporter.setStyleSheet(f"""
            QPushButton {{
                background-color: {COULEURS['accent_succes']};
                color: #1E1E2E;
                border: none;
                border-radius: 6px;
                font-size: 14px;
                font-weight: 700;
                padding: 8px 24px;
            }}
            QPushButton:hover {{ background-color: #88FFB8; }}
            QPushButton:disabled {{ background-color: {COULEURS['bordure']}; color: {COULEURS['texte_secondaire']}; }}
        """)
        self.btn_exporter.clicked.connect(self._lancer_export)
        row_btn.addWidget(self.btn_exporter)
        row_btn.addStretch()
        layout.addLayout(row_btn)

        # ── Barre de progression ─────────────────────────────
        self.progress = QProgressBar()
        self.progress.setRange(0, 0)  # indéterminé
        self.progress.setVisible(False)
        self.progress.setFixedHeight(8)
        self.progress.setStyleSheet(f"""
            QProgressBar {{
                border: none;
                background-color: {COULEURS['bg_carte']};
                border-radius: 4px;
            }}
            QProgressBar::chunk {{
                background-color: {COULEURS['accent_succes']};
                border-radius: 4px;
            }}
        """)
        layout.addWidget(self.progress)

        # ── Console logs ─────────────────────────────────────
        self.console = QTextEdit()
        self.console.setReadOnly(True)
        self.console.setFixedHeight(200)
        self.console.setStyleSheet(f"""
            QTextEdit {{
                background-color: {COULEURS['bg_carte']};
                color: {COULEURS['texte']};
                border: 1px solid {COULEURS['bordure']};
                border-radius: 6px;
                font-family: 'Consolas', monospace;
                font-size: 11px;
                padding: 8px;
            }}
        """)
        layout.addWidget(self.console)

        # ── Légende ──────────────────────────────────────────
        grp_legende = QGroupBox("Légende — Feuille 1")
        grp_legende.setStyleSheet(grp_params.styleSheet())
        lay_leg = QHBoxLayout(grp_legende)
        legendes = [
            ("Abs Total", "Nombre de jours ouvrés d'absence sur l'année"),
            ("Abs AM",    "Absences tombant sur un jour de cycle AM"),
            ("Jours AM",  "Jours travaillés en cycle AM (réels + hypothétiques)"),
            ("AM Nets",   "Jours AM - Abs AM"),
        ]
        for titre_leg, desc in legendes:
            col = QVBoxLayout()
            lbl_t = QLabel(f"<b>{titre_leg}</b>")
            lbl_t.setStyleSheet(f"color: {COULEURS['accent']}; font-size: 11px;")
            lbl_t.setTextFormat(Qt.TextFormat.RichText)
            lbl_d = QLabel(desc)
            lbl_d.setStyleSheet(f"color: {COULEURS['texte_secondaire']}; font-size: 10px;")
            lbl_d.setWordWrap(True)
            col.addWidget(lbl_t)
            col.addWidget(lbl_d)
            lay_leg.addLayout(col)
        layout.addWidget(grp_legende)

        layout.addStretch()

    def _lancer_export(self):
        # Choisir le fichier de sortie
        chemin, _ = QFileDialog.getSaveFileName(
            self, "Enregistrer l'export Excel",
            f"synthese_AM_{_date.today().strftime('%Y%m%d')}.xlsx",
            "Excel (*.xlsx)"
        )
        if not chemin:
            return

        self.console.clear()
        self.btn_exporter.setEnabled(False)
        self.progress.setVisible(True)

        qd = self.date_debut.date()
        qf = self.date_fin.date()
        d_debut = _date(qd.year(), qd.month(), qd.day())
        d_fin   = _date(qf.year(), qf.month(), qf.day())

        self._worker = WorkerExport(
            chemin_sortie    = chemin,
            d_debut          = d_debut,
            d_fin            = d_fin,
            inclure_feuille2 = self.check_feuille2.isChecked(),
        )
        self._worker.log_signal.connect(self.console.append)
        self._worker.fini_signal.connect(self._on_fini)
        self._worker.erreur_signal.connect(self._on_erreur)
        self._worker.start()

    def _on_fini(self, chemin: str):
        self.btn_exporter.setEnabled(True)
        self.progress.setVisible(False)
        try:
            self.window().status.showMessage(
                f"✅  Export généré : {os.path.basename(chemin)}", 5000
            )
        except Exception:
            pass
        rep = QMessageBox.information(
            self, "Export terminé",
            f"✅  Fichier Excel généré avec succès !\n\n{chemin}\n\n"
            "Voulez-vous ouvrir le fichier ?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if rep == QMessageBox.StandardButton.Yes:
            import subprocess, sys
            if sys.platform == "win32":
                os.startfile(chemin)
            elif sys.platform == "darwin":
                subprocess.Popen(["open", chemin])
            else:
                subprocess.Popen(["xdg-open", chemin])

    def _on_erreur(self, msg: str):
        self.btn_exporter.setEnabled(True)
        self.progress.setVisible(False)
        self.console.append(f"\n❌  ERREUR : {msg}")
        QMessageBox.critical(self, "Erreur lors de l'export", msg[:500])