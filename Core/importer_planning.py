"""
importer_planning.py
====================
Importe les données réelles dans planning_historique.json depuis :
  - Source 1 : Excel annuels Fabrication (2021→2024)
  - Source 2 : PDF ADP mensuel 2025 (Fabrication)
  - Source 3 : Excel ADP hebdomadaire tous départements

Format de sortie : planning_historique.json
{
  "NOM PRÉNOM|ID": {
    "semaines": {
      "S01_2024": {"cycle": "AM", "hypothetique": false, "source": "excel_fab"},
      ...
    },
    "jours": {
      "2025-08-04": {"cycle": "AM", "hypothetique": false, "source": "pdf_adp"},
      ...
    }
  }
}

Règles fondamentales :
  - Données réelles → hypothetique: false (priorité absolue sur les hypothétiques)
  - Cellule vide Excel → cycle: null, hypothetique: false (absence exceptionnelle)
  - Cycles null ne sont PAS reproduits dans les hypothétiques
  - Noms normalisés en MAJUSCULES, sans accents, suffixes *sst supprimés
  - Champ source : "excel_fab" | "pdf_adp" | "excel_adp" | "manuel"
  - Doublon (deux imports sur même jour) → signalé via callback_doublon, pas d'écrasement auto
"""

import json
import os
import re
from datetime import date, datetime, timedelta
from collections import defaultdict

# ─────────────────────────────────────────────
# Dépendances optionnelles (gérées proprement)
# ─────────────────────────────────────────────
try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

try:
    import pdfplumber
    PDFPLUMBER_OK = True
except ImportError:
    PDFPLUMBER_OK = False


# ═══════════════════════════════════════════════════════════════════════
# NORMALISATION DES NOMS
# ═══════════════════════════════════════════════════════════════════════

_ACCENTS_MAP = str.maketrans(
    "éèêëàâùûîïôçÉÈÊËÀÂÙÛÎÏÔÇ",
    "eeeéaauuiioceeeéaauuiioc".replace("é", "e")
)
# Table correcte construite manuellement
_ACCENT_PAIRS = [
    ('é','e'),('è','e'),('ê','e'),('ë','e'),
    ('à','a'),('â','a'),('ù','u'),('û','u'),
    ('î','i'),('ï','i'),('ô','o'),('ç','c'),
    ('É','E'),('È','E'),('Ê','E'),('Ë','E'),
    ('À','A'),('Â','A'),('Ù','U'),('Û','U'),
    ('Î','I'),('Ï','I'),('Ô','O'),('Ç','C'),
    ('ü','u'),('Ü','U'),
]

def supprimer_accents(s: str) -> str:
    for src, dst in _ACCENT_PAIRS:
        s = s.replace(src, dst)
    return s

def normaliser_nom(nom: str) -> str:
    """
    Normalise un nom d'employé :
    - Supprime espaces trailing/leading
    - Supprime suffixes *sst / * sst
    - Met en MAJUSCULES
    - Normalise espaces multiples
    Exemples :
      "FROISSART Pascal "   → "FROISSART PASCAL"
      "CIZEAU Benoit *sst"  → "CIZEAU BENOIT"
      "MERCEY Alexis * sst" → "MERCEY ALEXIS"
      "HEMERY-\nDUFOUG ERIC"→ "HEMERY- DUFOUG ERIC" (tiret conservé)
    """
    if not nom:
        return ''
    nom = str(nom).strip()
    # Supprimer suffixes *sst
    nom = re.sub(r'\s*\*?\s*sst\s*$', '', nom, flags=re.IGNORECASE)
    # Normaliser retours à la ligne et espaces multiples
    nom = re.sub(r'[\r\n]+', ' ', nom)
    nom = re.sub(r'\s+', ' ', nom)
    nom = nom.strip().upper()
    return nom


# ═══════════════════════════════════════════════════════════════════════
# NORMALISATION DES VALEURS DE CYCLE (Source 1)
# ═══════════════════════════════════════════════════════════════════════

# Mapping complet depuis le fichier Excel annuel
_MAP_CYCLE_EXCEL = {
    'M':   'M',
    'APM': 'AM',   # APM → AM
    'AM':  'AM',
    'N':   'N',
    'WE':  'WE',
    'J':   'J',    # apparu en 2024
}

def normaliser_cycle_excel(valeur) -> str | None:
    """
    Normalise une valeur de cycle depuis Excel annuel.
    Retourne None pour cellule vide (absence exceptionnelle).
    """
    if valeur is None:
        return None
    s = str(valeur).strip()
    if s == '':
        return None
    return _MAP_CYCLE_EXCEL.get(s.upper(), s.upper())


# ═══════════════════════════════════════════════════════════════════════
# CLÉS DE SEMAINE ET DE JOUR
# ═══════════════════════════════════════════════════════════════════════

def cle_semaine(num_sem: int, annee: int) -> str:
    """Ex: (1, 2023) → 'S01_2023'"""
    return f"S{num_sem:02d}_{annee}"

def cle_jour(annee: int, mois: int, jour: int) -> str:
    """Ex: (2025, 8, 4) → '2025-08-04'"""
    return f"{annee:04d}-{mois:02d}-{jour:02d}"


# ═══════════════════════════════════════════════════════════════════════
# CHARGEMENT / SAUVEGARDE DU PLANNING HISTORIQUE
# ═══════════════════════════════════════════════════════════════════════

def charger_planning(chemin: str) -> dict:
    """Charge planning_historique.json ou retourne un dict vide."""
    if not os.path.exists(chemin):
        return {}
    with open(chemin, 'r', encoding='utf-8') as f:
        return json.load(f)

def sauvegarder_planning(chemin: str, planning: dict) -> None:
    """Sauvegarde planning_historique.json (tri des clés pour lisibilité)."""
    # Trier les entrées
    planning_trie = {}
    for emp, data in sorted(planning.items()):
        planning_trie[emp] = {
            "semaines": dict(sorted(data.get("semaines", {}).items())),
            "jours":    dict(sorted(data.get("jours", {}).items())),
        }
    with open(chemin, 'w', encoding='utf-8') as f:
        json.dump(planning_trie, f, ensure_ascii=False, indent=2)


def _get_ou_creer_employe(planning: dict, cle_emp: str) -> dict:
    """Retourne ou crée l'entrée d'un employé dans le planning."""
    if cle_emp not in planning:
        planning[cle_emp] = {"semaines": {}, "jours": {}}
    return planning[cle_emp]


# ═══════════════════════════════════════════════════════════════════════
# BLACKLIST
# ═══════════════════════════════════════════════════════════════════════

def charger_blacklist(chemin: str) -> set:
    """Charge employes_blacklist.json → ensemble de noms normalisés."""
    if not os.path.exists(chemin):
        return set()
    with open(chemin, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return {normaliser_nom(n) for n in data}


# ═══════════════════════════════════════════════════════════════════════
# MATCHING DES NOMS D'EMPLOYÉS
# ═══════════════════════════════════════════════════════════════════════

def distance_levenshtein(a: str, b: str) -> int:
    """Distance de Levenshtein entre deux chaînes."""
    if a == b:
        return 0
    la, lb = len(a), len(b)
    if la == 0: return lb
    if lb == 0: return la
    prev = list(range(lb + 1))
    for i in range(1, la + 1):
        curr = [i] + [0] * lb
        for j in range(1, lb + 1):
            ins = prev[j] + 1
            dlt = curr[j-1] + 1
            sub = prev[j-1] + (0 if a[i-1] == b[j-1] else 1)
            curr[j] = min(ins, dlt, sub)
        prev = curr
    return prev[lb]

def trouver_employe(
    nom_normalise: str,
    employes_connus: dict,
    blacklist: set,
    seuil_fuzzy: int = 2
) -> tuple[str | None, str]:
    """
    Cherche un employé dans employes_connus par son nom normalisé.
    employes_connus : {cle_emp: {"nom": str, ...}} ou {"NOM PRÉNOM|ID": ...}

    Retourne (cle_trouvee, statut) où statut est :
      "exact"     : correspondance parfaite
      "fuzzy"     : correspondance fuzzy (Levenshtein ≤ seuil_fuzzy)
      "blacklist" : nom blacklisté
      "inconnu"   : non trouvé
    """
    if nom_normalise in blacklist:
        return None, "blacklist"

    # Index nom → clé (construction à la demande)
    for cle, info in employes_connus.items():
        # La clé est "NOM PRÉNOM|ID" → extraire le nom
        nom_emp = cle.split('|')[0] if '|' in cle else cle
        if nom_emp == nom_normalise:
            return cle, "exact"

    # Fuzzy matching
    meilleur_cle = None
    meilleure_dist = seuil_fuzzy + 1
    for cle in employes_connus:
        nom_emp = cle.split('|')[0] if '|' in cle else cle
        # Comparer sans accents
        d = distance_levenshtein(
            supprimer_accents(nom_normalise),
            supprimer_accents(nom_emp)
        )
        if d <= seuil_fuzzy and d < meilleure_dist:
            meilleure_dist = d
            meilleur_cle = cle

    if meilleur_cle:
        return meilleur_cle, "fuzzy"

    return None, "inconnu"


# ═══════════════════════════════════════════════════════════════════════
# SOURCE 1 — EXCEL ANNUELS FABRICATION (2021→2024)
# ═══════════════════════════════════════════════════════════════════════

def _est_ligne_employe(row: tuple) -> bool:
    """Vérifie que la ligne est une ligne d'employé valide (col A non vide, col B non None)."""
    col_a = str(row[0] or '').strip()
    col_b = row[1]
    return col_a != '' and col_b is not None

def importer_excel_annuel(
    chemin_xlsx: str,
    chemin_planning: str,
    chemin_employes: str,
    chemin_blacklist: str = '',
    callback_log=None,
    callback_non_reconnu=None,
) -> dict:
    """
    Importe un fichier Excel annuel Fabrication dans planning_historique.json.

    Structure du fichier :
      L1 : en-tête → col A = année, col B = "Nom / Prenom", col C+ = S1, S2, ... S52
      L2+ : une ligne par employé
        col A = type contrat (CDI, CDD, intérim...)
        col B = nom format "NOM Prénom" → normalisé en "NOM PRÉNOM"
        col C→fin = cycle par semaine (M, APM, N, WE, J, None)

    Retourne un dict de résumé : {
      "employes_importes": int,
      "semaines_ecrites": int,
      "non_reconnus": list[str],
      "avertissements": list[str],
    }
    """
    def log(msg):
        if callback_log:
            callback_log(msg)

    if not OPENPYXL_OK:
        raise ImportError("openpyxl est requis pour importer les Excel annuels. pip install openpyxl")

    log(f"📖 Lecture : {os.path.basename(chemin_xlsx)}")

    # Chargement des données existantes
    planning = charger_planning(chemin_planning)
    blacklist = charger_blacklist(chemin_blacklist) if chemin_blacklist else set()

    # Chargement de la liste des employés connus
    employes_connus = {}
    if os.path.exists(chemin_employes):
        with open(chemin_employes, 'r', encoding='utf-8') as f:
            employes_connus = json.load(f)

    # Lecture du classeur
    wb = openpyxl.load_workbook(chemin_xlsx, read_only=True, data_only=True)
    ws = wb.worksheets[0]  # Première feuille uniquement

    # Lecture de l'en-tête (L1)
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    annee = int(header[0]) if header[0] else None
    if not annee:
        wb.close()
        raise ValueError(f"Impossible de lire l'année en col A de la L1 : {chemin_xlsx}")

    # Colonnes semaines : col C+ (index 2+)
    colonnes_semaines = []
    for i, val in enumerate(header[2:], start=3):
        if val and str(val).startswith('S'):
            try:
                num_sem = int(str(val)[1:])
                colonnes_semaines.append((i, num_sem, cle_semaine(num_sem, annee)))
            except ValueError:
                pass

    log(f"   Année : {annee} | {len(colonnes_semaines)} semaines ({colonnes_semaines[0][2]} → {colonnes_semaines[-1][2]})")

    # Compteurs
    nb_employes = 0
    nb_semaines = 0
    non_reconnus = []
    avertissements = []

    # Lecture des lignes d'employés (L2+)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not _est_ligne_employe(row):
            continue

        nom_brut = str(row[1]).strip() if row[1] else ''
        if not nom_brut:
            continue

        nom_norm = normaliser_nom(nom_brut)
        if not nom_norm:
            continue

        # Matching employé
        cle_emp, statut = trouver_employe(nom_norm, employes_connus, blacklist)

        if statut == "blacklist":
            log(f"   ⬛ Ignoré (blacklist) : {nom_norm}")
            continue

        if statut == "inconnu":
            log(f"   ❓ Non reconnu : {nom_norm}")
            non_reconnus.append(nom_norm)
            if callback_non_reconnu:
                callback_non_reconnu(nom_norm, "excel_annuel", annee)
            # On importe quand même avec le nom comme clé (sans ID)
            cle_emp = f"{nom_norm}|"

        elif statut == "fuzzy":
            msg = f"   ⚠️  Fuzzy match : '{nom_norm}' → '{cle_emp}'"
            log(msg)
            avertissements.append(msg)

        # Écriture des cycles dans le planning
        emp_data = _get_ou_creer_employe(planning, cle_emp)
        semaines_dict = emp_data["semaines"]

        nb_semaines_emp = 0
        for col_idx, num_sem, cle_sem in colonnes_semaines:
            # col_idx est 1-based dans l'en-tête, mais row est un tuple 0-based
            valeur_brute = row[col_idx - 1] if col_idx - 1 < len(row) else None
            cycle = normaliser_cycle_excel(valeur_brute)

            # Ne pas écraser une donnée réelle existante (hypothetique: false) par une autre réelle
            # Priorité : données réelles > hypothétiques
            if cle_sem in semaines_dict and not semaines_dict[cle_sem].get("hypothetique", True):
                # Donnée réelle déjà présente → on la conserve (import = idempotent)
                pass
            else:
                semaines_dict[cle_sem] = {
                    "cycle": cycle,
                    "hypothetique": False,
                    "source": "excel_fab",
                }
                nb_semaines_emp += 1

        nb_employes += 1
        nb_semaines += nb_semaines_emp
        log(f"   ✅ {nom_norm} ({nb_semaines_emp} semaines)")

    wb.close()

    # Sauvegarde
    sauvegarder_planning(chemin_planning, planning)

    log(f"\n📊 Résumé import Excel {annee} :")
    log(f"   {nb_employes} employé(s) importé(s)")
    log(f"   {nb_semaines} entrée(s) de semaine écrite(s)")
    if non_reconnus:
        log(f"   ⚠️  {len(non_reconnus)} non reconnu(s) : {', '.join(non_reconnus[:5])}")

    return {
        "employes_importes": nb_employes,
        "semaines_ecrites": nb_semaines,
        "non_reconnus": non_reconnus,
        "avertissements": avertissements,
        "annee": annee,
    }


# ═══════════════════════════════════════════════════════════════════════
# SOURCE 2 — PDF ADP MENSUEL 2025 (FABRICATION)
# ═══════════════════════════════════════════════════════════════════════

# Mapping des mots du PDF → cycle
_MOTS_TYPE_PDF = {"Posté", "Repos", "Administratif"}
_MOTS_QUAL_PDF = {"AM", "Matin", "Nuit", "jour", "Week", "Journée", "NON", "end", "semaine"}

_PATTERN_JOUR_ENTETE = re.compile(r'^(Lun|Mar|Mer|Jeu|Ven|Sam|Dim)\.(\d{1,2})$')
_PATTERN_NOM_MAJUSCULE = re.compile(r'^[A-ZÀÂÉÈÊËÎÏÔÙÛÜÇŒÆ\-]+$')

def _est_bande_type(mots_bande: list) -> bool:
    """Vrai si la bande contient 'Posté', 'Repos' ou 'Administratif' comme mot exact."""
    return any(w['text'] in _MOTS_TYPE_PDF for w in mots_bande)

def _est_bande_qual(mots_bande: list) -> bool:
    """Vrai si la bande contient un qualificateur de cycle comme mot exact."""
    return any(w['text'] in _MOTS_QUAL_PDF for w in mots_bande)

def _combiner_cycle_pdf(mot_type: str | None, mot_qual: str | None) -> str | None:
    """
    Détermine le cycle depuis le mot TYPE (Posté/Repos/Administratif)
    et le mot QUAL (AM/Matin/Nuit/jour/Week/Journée/NON/end/semaine).
    """
    l1 = (mot_type or '').lower()
    l2 = (mot_qual or '').lower()

    if 'repos' in l1 and ('jour' in l2 or 'semaine' in l2): return 'R'
    if 'repos' in l1: return 'R'
    if 'posté' in l1 and l2 == 'am': return 'AM'
    if 'posté' in l1 and 'matin' in l2: return 'M'
    if 'posté' in l1 and 'nuit' in l2: return 'N'
    if 'posté' in l1 and ('week' in l2 or 'end' in l2): return 'WE'
    if 'posté' in l1 and ('journée' in l2 or 'journee' in l2): return 'J'
    if 'posté' in l1 and 'jour' in l2: return 'J'
    if 'administratif' in l1: return 'J'
    if 'posté' in l1 and 'non' in l2: return 'N'   # "Posté Nuit NON EL"
    return None

def _extraire_page_pdf(page, callback_log=None) -> dict:
    """
    Extrait les cycles de tous les employés d'une page du PDF ADP.
    Retourne : {nom_normalise: {num_jour: cycle}}
    """
    def log(msg):
        if callback_log:
            callback_log(msg)

    words = page.extract_words(x_tolerance=3, y_tolerance=3)

    # 1. Colonnes jours (depuis l'en-tête "Ven.1 Sam.2 ...")
    colonnes_jours = {}
    for w in words:
        m = _PATTERN_JOUR_ENTETE.match(w['text'])
        if m:
            num_jour = int(m.group(2))
            colonnes_jours[num_jour] = (w['x0'] + w['x1']) / 2

    if not colonnes_jours:
        log("   ⚠️  Pas de colonnes jours détectées sur cette page")
        return {}

    nb_jours = len(colonnes_jours)
    col_width = 37.0  # largeur approximative d'une colonne
    if nb_jours >= 2:
        jours_tries = sorted(colonnes_jours.items())
        col_width = (jours_tries[-1][1] - jours_tries[0][1]) / (nb_jours - 1)

    def x_to_jour(x: float) -> int | None:
        best, best_dist = None, 9999.0
        for j, cx in colonnes_jours.items():
            d = abs(cx - x)
            if d < best_dist:
                best_dist = d
                best = j
        return best if best_dist < col_width * 0.65 else None

    # 2. Grouper les mots par bande Y (arrondi à 4px)
    bandes: dict[int, list] = defaultdict(list)
    for w in words:
        band_y = round(w['top'] / 4) * 4
        bandes[band_y].append(w)
    bandes_ys = sorted(bandes.keys())

    def get_mots_par_jour(y_bande: int) -> dict:
        """Extrait le premier mot significatif par colonne de jour."""
        mots_j = {}
        for w in sorted(bandes[y_bande], key=lambda w: w['x0']):
            j = x_to_jour((w['x0'] + w['x1']) / 2)
            if j is not None and j not in mots_j:
                mots_j[j] = w['text']
        return mots_j

    # 3. Identifier les blocs d'employés
    #    Chaque employé est sur 1 ou 2 lignes en colonne x < 110
    employes_bruts: list[tuple[int, str]] = []
    for y in bandes_ys:
        mots_nom = [
            w for w in bandes[y]
            if w['x0'] < 110 and _PATTERN_NOM_MAJUSCULE.match(w['text'])
        ]
        if mots_nom:
            texte = ' '.join(w['text'] for w in sorted(mots_nom, key=lambda w: w['x0']))
            employes_bruts.append((y, texte))

    # Consolider NOM + PRÉNOM (2 lignes consécutives)
    employes: list[tuple[int, int, str]] = []  # (y_ligne1, y_ligne2, nom)
    i = 0
    while i < len(employes_bruts):
        y1, n1 = employes_bruts[i]
        if (i + 1 < len(employes_bruts)
                and employes_bruts[i + 1][0] - y1 < 15):
            y2, n2 = employes_bruts[i + 1]
            employes.append((y1, y2, f"{n1} {n2}"))
            i += 2
        else:
            employes.append((y1, y1, n1))
            i += 1

    # 4. Pour chaque employé, trouver les bandes TYPE et QUAL
    resultats: dict[str, dict[int, str]] = {}

    for y_nom1, y_nom2, nom_brut in employes:
        # Chercher la première paire (TYPE, QUAL) dans la fenêtre [-30, -4] depuis y_nom2
        y_type = None
        y_qual = None

        for y in bandes_ys:
            delta = y - y_nom2
            if -30 <= delta <= -4 and _est_bande_type(bandes[y]):
                # Chercher une bande QUAL dans les 4-12px suivants
                for y2 in bandes_ys:
                    d2 = y2 - y
                    if 4 <= d2 <= 12 and _est_bande_qual(bandes[y2]):
                        y_type = y
                        y_qual = y2
                        break
                if y_type is not None:
                    break  # Première paire valide trouvée

        if y_type is None:
            log(f"   ⚠️  {nom_brut} : bande poste non trouvée")
            continue

        mots_type = get_mots_par_jour(y_type)
        mots_qual = get_mots_par_jour(y_qual) if y_qual else {}

        cycles: dict[int, str] = {}
        for j in range(1, 32):
            c = _combiner_cycle_pdf(mots_type.get(j), mots_qual.get(j))
            if c is not None:
                cycles[j] = c

        nom_norm = normaliser_nom(nom_brut)
        resultats[nom_norm] = cycles

    return resultats


def _parser_entete_pdf(page) -> tuple[int, int] | None:
    """
    Extrait (mois, annee) depuis la ligne d'en-tête du PDF :
    "Etat des congés et évènements du 01/08/2025 au 31/08/2025"
    Retourne None si non trouvé.
    """
    texte = page.extract_text() or ''
    m = re.search(r'du\s+\d{2}/(\d{2})/(\d{4})\s+au', texte)
    if m:
        mois = int(m.group(1))
        annee = int(m.group(2))
        return mois, annee
    return None

def importer_pdf_adp(
    chemin_pdf: str,
    chemin_planning: str,
    chemin_employes: str,
    chemin_blacklist: str = '',
    callback_log=None,
    callback_non_reconnu=None,
    callback_doublon=None,
) -> dict:
    """
    Importe un PDF ADP mensuel dans planning_historique.json.

    Le PDF contient les postes théoriques jour par jour pour tous les employés
    de la population FABRICATION.

    Règles :
      - Le poste théorique est stocké MÊME SI l'employé est absent ce jour
      - Les absences (CP, JS, ZV...) sont gérées par absences.py → non traitées ici
      - Données réelles : hypothetique: false

    Retourne un dict de résumé.
    """
    def log(msg):
        if callback_log:
            callback_log(msg)

    if not PDFPLUMBER_OK:
        raise ImportError("pdfplumber est requis pour importer les PDF ADP. pip install pdfplumber")

    log(f"📖 Lecture PDF : {os.path.basename(chemin_pdf)}")

    planning = charger_planning(chemin_planning)
    blacklist = charger_blacklist(chemin_blacklist) if chemin_blacklist else set()

    employes_connus = {}
    if os.path.exists(chemin_employes):
        with open(chemin_employes, 'r', encoding='utf-8') as f:
            employes_connus = json.load(f)

    nb_employes = 0
    nb_jours_ecrits = 0
    non_reconnus = []
    avertissements = []

    mois_annee = None  # (mois, annee)

    with pdfplumber.open(chemin_pdf) as pdf:
        for page_idx, page in enumerate(pdf.pages):
            # Extraire mois/année depuis la première page
            if page_idx == 0:
                mois_annee = _parser_entete_pdf(page)
                if mois_annee:
                    mois, annee = mois_annee
                    log(f"   Période : {mois:02d}/{annee}")
                else:
                    log("   ⚠️  Impossible de détecter la période (mois/année)")
                    annee, mois = 2025, 1  # fallback

            log(f"   Page {page_idx + 1}...")
            cycles_page = _extraire_page_pdf(page, callback_log)

            for nom_norm, cycles_jours in cycles_page.items():
                # Matching employé
                cle_emp, statut = trouver_employe(nom_norm, employes_connus, blacklist)

                if statut == "blacklist":
                    log(f"      ⬛ Ignoré (blacklist) : {nom_norm}")
                    continue

                if statut == "inconnu":
                    log(f"      ❓ Non reconnu : {nom_norm}")
                    non_reconnus.append(nom_norm)
                    if callback_non_reconnu:
                        callback_non_reconnu(nom_norm, "pdf_adp", mois_annee)
                    cle_emp = f"{nom_norm}|"

                elif statut == "fuzzy":
                    msg = f"      ⚠️  Fuzzy : '{nom_norm}' → '{cle_emp}'"
                    log(msg)
                    avertissements.append(msg)

                emp_data = _get_ou_creer_employe(planning, cle_emp)
                jours_dict = emp_data["jours"]

                nb_jours_emp = 0
                for num_jour, cycle in cycles_jours.items():
                    cle_j = cle_jour(annee, mois, num_jour)
                    # Doublon : jour déjà en hypothetique: false → signaler, ne pas écraser
                    if cle_j in jours_dict and not jours_dict[cle_j].get("hypothetique", True):
                        existant = jours_dict[cle_j]
                        if existant.get("cycle") != cycle and callback_doublon:
                            callback_doublon(cle_emp, cle_j, existant, {
                                "cycle": cycle,
                                "hypothetique": False,
                                "source": "pdf_adp",
                            })
                        # Sans confirmation explicite on conserve l'existant
                    else:
                        jours_dict[cle_j] = {
                            "cycle": cycle,
                            "hypothetique": False,
                            "source": "pdf_adp",
                        }
                        nb_jours_emp += 1

                if nb_jours_emp > 0:
                    nb_employes += 1
                    nb_jours_ecrits += nb_jours_emp
                    log(f"      ✅ {nom_norm} ({nb_jours_emp} jours)")

    sauvegarder_planning(chemin_planning, planning)

    log(f"\n📊 Résumé import PDF ADP :")
    log(f"   {nb_employes} employé(s) importé(s)")
    log(f"   {nb_jours_ecrits} entrée(s) de jour écrite(s)")
    if non_reconnus:
        log(f"   ⚠️  {len(non_reconnus)} non reconnu(s) : {', '.join(non_reconnus[:5])}")

    return {
        "employes_importes": nb_employes,
        "jours_ecrits": nb_jours_ecrits,
        "non_reconnus": non_reconnus,
        "avertissements": avertissements,
        "periode": mois_annee,
    }


# ═══════════════════════════════════════════════════════════════════════
# IMPORT EN LOT
# ═══════════════════════════════════════════════════════════════════════

def importer_tous_excel(
    dossier: str,
    chemin_planning: str,
    chemin_employes: str,
    chemin_blacklist: str = '',
    callback_log=None,
    callback_non_reconnu=None,
) -> list[dict]:
    """
    Importe tous les fichiers Excel annuels (2021.xlsx, 2022.xlsx...) d'un dossier.
    Retourne la liste des résumés d'import.
    """
    def log(msg):
        if callback_log:
            callback_log(msg)

    resultats = []
    fichiers = sorted([
        f for f in os.listdir(dossier)
        if f.endswith('.xlsx') and re.match(r'^\d{4}\.xlsx$', f)
    ])

    if not fichiers:
        log("⚠️  Aucun fichier Excel annuel trouvé (format attendu : 2021.xlsx, 2022.xlsx...)")
        return resultats

    for fichier in fichiers:
        chemin = os.path.join(dossier, fichier)
        log(f"\n{'='*50}")
        log(f"🗂️  Import : {fichier}")
        log('='*50)
        try:
            r = importer_excel_annuel(
                chemin_xlsx=chemin,
                chemin_planning=chemin_planning,
                chemin_employes=chemin_employes,
                chemin_blacklist=chemin_blacklist,
                callback_log=callback_log,
                callback_non_reconnu=callback_non_reconnu,
            )
            resultats.append(r)
        except Exception as e:
            log(f"❌ Erreur lors de l'import de {fichier} : {e}")

    return resultats


def importer_tous_pdf(
    dossier: str,
    chemin_planning: str,
    chemin_employes: str,
    chemin_blacklist: str = '',
    callback_log=None,
    callback_non_reconnu=None,
) -> list[dict]:
    """
    Importe tous les PDF ADP d'un dossier.
    Retourne la liste des résumés d'import.
    """
    def log(msg):
        if callback_log:
            callback_log(msg)

    resultats = []
    fichiers = sorted([
        f for f in os.listdir(dossier)
        if f.lower().endswith('.pdf')
    ])

    if not fichiers:
        log("⚠️  Aucun fichier PDF trouvé dans le dossier")
        return resultats

    for fichier in fichiers:
        chemin = os.path.join(dossier, fichier)
        log(f"\n{'='*50}")
        log(f"🗂️  Import PDF : {fichier}")
        log('='*50)
        try:
            r = importer_pdf_adp(
                chemin_pdf=chemin,
                chemin_planning=chemin_planning,
                chemin_employes=chemin_employes,
                chemin_blacklist=chemin_blacklist,
                callback_log=callback_log,
                callback_non_reconnu=callback_non_reconnu,
            )
            resultats.append(r)
        except Exception as e:
            log(f"❌ Erreur lors de l'import de {fichier} : {e}")

    return resultats

# ═══════════════════════════════════════════════════════════════════════
# SOURCE 3 — EXCEL ADP HEBDOMADAIRE (TOUS DÉPARTEMENTS)
# ═══════════════════════════════════════════════════════════════════════

def _normaliser_id_adp(id_val) -> str:
    """
    Normalise un ID depuis l'Excel ADP hebdo.
    L'Excel stocke des entiers sans zéros (926, 90039…).
    On normalise en str(int) puis zfill(8) pour matcher les clés employes_contrats.
    Ex: 926 → "00000926", 90039 → "00090039"
    """
    if id_val is None:
        return ''
    try:
        return str(int(float(str(id_val)))).zfill(8)
    except (ValueError, TypeError):
        return str(id_val).strip()


def _trouver_employe_par_id(
    id_normalise: str,
    employes_connus: dict,
) -> str | None:
    """
    Cherche un employé dans employes_connus par son ID normalisé (zfill 8).
    Retourne la clé trouvée ou None.
    """
    if not id_normalise:
        return None
    for cle in employes_connus:
        if '|' in cle:
            id_part = cle.split('|')[1]
            if id_part == id_normalise:
                return cle
    return None


def _extraire_dates_l2(row_l2: tuple) -> list:
    """
    Extrait les dates de la ligne L2 de l'Excel ADP hebdo.
    Robuste au décalage S43 (dates commençant en col B au lieu de C).
    Retourne une liste de (col_index_0based, date) pour les colonnes C+ (index ≥ 2).
    """
    dates = []
    for i, val in enumerate(row_l2):
        if val is None:
            continue
        d = None
        if isinstance(val, datetime):
            d = val.date()
        elif isinstance(val, date) and not isinstance(val, datetime):
            d = val
        elif isinstance(val, str):
            try:
                d = datetime.fromisoformat(val.split()[0]).date()
            except ValueError:
                pass
        if d is not None and i >= 1:  # au minimum col B (index 1)
            dates.append((i, d))
    return dates


def importer_excel_adp_hebdo(
    chemin_xlsx: str,
    chemin_planning: str,
    chemin_employes: str,
    chemin_blacklist: str = '',
    callback_log=None,
    callback_non_reconnu=None,
    callback_doublon=None,
) -> dict:
    """
    Importe un fichier Excel ADP hebdomadaire dans planning_historique.json.

    Structure du fichier (feuille nommée ex: 'S41_06-10-2025') :
      L1 : Nom / Prénom | ID | Lundi | Mardi | Mercredi | Jeudi | Vendredi | Samedi | Dimanche
      L2 : (vide)       | (vide) | date  | date  | date  | date  | date    | date   | date
      L3+: NOM PRÉNOM   | ID_int | cycle | cycle | cycle | cycle | cycle   | cycle  | cycle

    Règles :
      - Matching ID prioritaire : str(id).zfill(8) → chercher dans les clés employes_contrats
      - Fuzzy Levenshtein ≤ 2 sur le nom en fallback si ID non trouvé
      - Doublon (jour déjà hypothetique:false) → callback_doublon, pas d'écrasement auto
      - source: "excel_adp", hypothetique: false
      - Idempotent : relancer ne duplique pas

    Retourne un dict de résumé.
    """
    def log(msg):
        if callback_log:
            callback_log(msg)

    if not OPENPYXL_OK:
        raise ImportError("openpyxl est requis. pip install openpyxl")

    log(f"📖 Lecture : {os.path.basename(chemin_xlsx)}")

    planning = charger_planning(chemin_planning)
    blacklist = charger_blacklist(chemin_blacklist) if chemin_blacklist else set()

    employes_connus = {}
    if os.path.exists(chemin_employes):
        with open(chemin_employes, 'r', encoding='utf-8') as f:
            employes_connus = json.load(f)

    wb = openpyxl.load_workbook(chemin_xlsx, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    nom_feuille = wb.sheetnames[0]
    log(f"   Feuille : {nom_feuille}")

    # Extraire le numéro de semaine depuis le nom de feuille (ex: S41_06-10-2025)
    m_sem = re.match(r'^S(\d+)_', nom_feuille)
    num_semaine = int(m_sem.group(1)) if m_sem else None

    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if len(rows) < 3:
        log("   ⚠️  Fichier vide ou trop court")
        return {"employes_importes": 0, "jours_ecrits": 0,
                "non_reconnus": [], "avertissements": [], "semaine": num_semaine}

    l2 = rows[1]

    # Extraire les dates de L2 (robuste au décalage de colonne)
    dates_cols = _extraire_dates_l2(l2)
    # On ne garde que les colonnes ≥ 2 (col C+) sauf si décalage S43
    # En pratique on prend toutes les dates trouvées en col index ≥ 1
    # et on les associe dans l'ordre aux jours de la semaine
    if not dates_cols:
        log("   ⚠️  Aucune date trouvée en L2")
        return {"employes_importes": 0, "jours_ecrits": 0,
                "non_reconnus": [], "avertissements": [], "semaine": num_semaine}

    log(f"   Dates : {[str(d) for _, d in dates_cols]}")

    # Compteurs
    nb_employes = 0
    nb_jours_ecrits = 0
    non_reconnus = []
    avertissements = []

    # Traiter L3+
    for row in rows[2:]:
        nom_brut = row[0]
        id_brut  = row[1]

        if nom_brut is None:
            continue

        nom_norm = normaliser_nom(str(nom_brut))
        if not nom_norm:
            continue

        # --- Matching : ID prioritaire, fuzzy fallback ---
        id_norm = _normaliser_id_adp(id_brut)
        cle_emp = _trouver_employe_par_id(id_norm, employes_connus)
        statut = "exact" if cle_emp else None

        if not cle_emp:
            # Fallback fuzzy sur le nom
            cle_emp, statut = trouver_employe(nom_norm, employes_connus, blacklist)

        if statut == "blacklist":
            log(f"   ⬛ Ignoré (blacklist) : {nom_norm}")
            continue

        if statut is None or statut == "inconnu":
            log(f"   ❓ Non reconnu : {nom_norm} (ID={id_norm})")
            non_reconnus.append(nom_norm)
            if callback_non_reconnu:
                # 3ème arg = id_norm : permet au dialogue non-reconnus de pré-remplir le matricule
                callback_non_reconnu(nom_norm, "excel_adp", num_semaine, id_norm)
            # Créer une clé temporaire avec l'ID brut pour ne pas perdre la donnée
            cle_emp = f"{nom_norm}|{id_norm}"

        elif statut == "fuzzy":
            msg = f"   ⚠️  Fuzzy : '{nom_norm}' (ID={id_norm}) → '{cle_emp}'"
            log(msg)
            avertissements.append(msg)

        # Écriture des cycles jour par jour
        emp_data = _get_ou_creer_employe(planning, cle_emp)
        jours_dict = emp_data["jours"]

        nb_jours_emp = 0
        for col_idx, d in dates_cols:
            if col_idx >= len(row):
                continue
            val_cycle = row[col_idx]
            if val_cycle is None:
                continue
            cycle = str(val_cycle).strip().upper()
            if not cycle:
                continue
            # Ignorer les valeurs non reconnues comme cycles (matricules, textes parasites)
            _CYCLES_VALIDES = {'M', 'AM', 'N', 'J', 'WE', 'R'}
            if cycle not in _CYCLES_VALIDES:
                continue

            cle_j = d.strftime("%Y-%m-%d")
            entree_new = {
                "cycle": cycle,
                "hypothetique": False,
                "source": "excel_adp",
            }

            if cle_j in jours_dict and not jours_dict[cle_j].get("hypothetique", True):
                # Doublon réel → signaler si cycle différent
                existant = jours_dict[cle_j]
                if existant.get("cycle") != cycle and callback_doublon:
                    callback_doublon(cle_emp, cle_j, existant, entree_new)
                # Pas d'écrasement sans confirmation
            else:
                jours_dict[cle_j] = entree_new
                nb_jours_emp += 1

        if nb_jours_emp > 0:
            nb_employes += 1
            nb_jours_ecrits += nb_jours_emp
            log(f"   ✅ {nom_norm} ({nb_jours_emp} jours)")

    sauvegarder_planning(chemin_planning, planning)

    log(f"\n📊 Résumé import Excel ADP hebdo {nom_feuille} :")
    log(f"   {nb_employes} employé(s) importé(s)")
    log(f"   {nb_jours_ecrits} entrée(s) de jour écrite(s)")
    if non_reconnus:
        log(f"   ⚠️  {len(non_reconnus)} non reconnu(s) : {', '.join(non_reconnus[:5])}")

    return {
        "employes_importes": nb_employes,
        "jours_ecrits": nb_jours_ecrits,
        "non_reconnus": non_reconnus,
        "avertissements": avertissements,
        "semaine": num_semaine,
        "feuille": nom_feuille,
    }


def importer_tous_excel_adp_hebdo(
    dossier: str,
    chemin_planning: str,
    chemin_employes: str,
    chemin_blacklist: str = '',
    callback_log=None,
    callback_non_reconnu=None,
    callback_doublon=None,
) -> list[dict]:
    """
    Importe tous les fichiers Excel ADP hebdomadaires d'un dossier.
    Détecte les fichiers au nom commençant par S + chiffres (S41.xlsx, S41_06-10-2025.xlsx...).
    Retourne la liste des résumés d'import.
    """
    def log(msg):
        if callback_log:
            callback_log(msg)

    resultats = []
    fichiers = sorted([
        f for f in os.listdir(dossier)
        if f.lower().endswith('.xlsx') and re.match(r'^S\d+', f, re.IGNORECASE)
    ])

    if not fichiers:
        log("⚠️  Aucun fichier Excel ADP hebdo trouvé (format attendu : S41.xlsx, S41_06-10-2025.xlsx...)")
        return resultats

    for fichier in fichiers:
        chemin = os.path.join(dossier, fichier)
        log(f"\n{'='*50}")
        log(f"🗂️  Import hebdo : {fichier}")
        log('='*50)
        try:
            r = importer_excel_adp_hebdo(
                chemin_xlsx=chemin,
                chemin_planning=chemin_planning,
                chemin_employes=chemin_employes,
                chemin_blacklist=chemin_blacklist,
                callback_log=callback_log,
                callback_non_reconnu=callback_non_reconnu,
                callback_doublon=callback_doublon,
            )
            resultats.append(r)
        except Exception as e:
            log(f"❌ Erreur lors de l'import de {fichier} : {e}")

    return resultats


# ═══════════════════════════════════════════════════════════════════════
# POINT D'ENTRÉE CLI (test / debug)
# ═══════════════════════════════════════════════════════════════════════

if __name__ == '__main__':
    import sys

    def log(msg):
        print(msg)

    if len(sys.argv) < 2:
        print("Usage:")
        print("  python importer_planning.py excel  <fichier.xlsx> [planning.json] [employes.json]")
        print("  python importer_planning.py pdf    <fichier.pdf>  [planning.json] [employes.json]")
        print("  python importer_planning.py hebdo  <fichier.xlsx> [planning.json] [employes.json]")
        sys.exit(1)

    mode = sys.argv[1]
    chemin_source = sys.argv[2] if len(sys.argv) > 2 else ''
    chemin_planning = sys.argv[3] if len(sys.argv) > 3 else 'planning_historique.json'
    chemin_employes = sys.argv[4] if len(sys.argv) > 4 else 'employes_contrats.json'

    if mode == 'excel':
        r = importer_excel_annuel(
            chemin_xlsx=chemin_source,
            chemin_planning=chemin_planning,
            chemin_employes=chemin_employes,
            callback_log=log,
        )
        print(f"\n✅ Import terminé : {r}")

    elif mode == 'pdf':
        r = importer_pdf_adp(
            chemin_pdf=chemin_source,
            chemin_planning=chemin_planning,
            chemin_employes=chemin_employes,
            callback_log=log,
        )
        print(f"\n✅ Import terminé : {r}")

    elif mode == 'hebdo':
        r = importer_excel_adp_hebdo(
            chemin_xlsx=chemin_source,
            chemin_planning=chemin_planning,
            chemin_employes=chemin_employes,
            callback_log=log,
        )
        print(f"\n✅ Import terminé : {r}")

    else:
        print(f"Mode inconnu : {mode}")
        sys.exit(1)